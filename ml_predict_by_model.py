# 准备当天的实时数据
import datetime
import time

import mysql
import openpyxl
import tushare as ts
import constants
import find_stock

def take_third(elem):
    return elem[3]

def prepare_data_3yang1tiao():
    prepare_data_3yang1tiao_chuangyeban()
    prepare_data_3yang1tiao_hushen()

def prepare_data_3yang1tiao_chuangyeban():
    prepare_data_3yang1tiao_with_filter(constants.stock_filter_chuangyeban, constants.ml_sheet_name_predict)

def prepare_data_3yang1tiao_hushen():
    prepare_data_3yang1tiao_with_filter(constants.stock_filter_hushen, constants.ml_sheet_name_predict_hushen)

def prepare_data_3yang1tiao_with_filter(filter, sheetname):
    date = time.strftime('%Y-%m-%d')
    # date = '2023-02-15'
    now = datetime.date(*map(int, date.split('-')))
    str_yestoday, yestoday = find_stock.get_pre_trade_day(now)
    str_yestoday_filename = constants.get_date_str_for_filename(yestoday)
    filename = constants.report_dir + str_yestoday_filename + constants.filename_3yang_list
    # 遍历所有2进3标的，将所需的数据写入到excel表predict页中
    fp = open(filename, 'r')
    lines = fp.readlines()
    fp.close()

    filename = constants.ml_data_dir + constants.ml_excel_name_3yang1tiao
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]
    curr_row = 2

    str_yestoday_yestoday, yestoday_yestoday = find_stock.get_pre_trade_day(yestoday)
    str_yestoday_yestoday_yestoday, yestoday_yestoday_yestoday = find_stock.get_pre_trade_day(yestoday_yestoday)
    conn = mysql.connector.connect(user=constants.mysql_user, password=constants.mysql_password,
                                   database=constants.mysql_database_name)
    cursor = conn.cursor()
    df = ts.get_hist_data('sh', start=date, end=date)
    change_index_sz = df.p_change[0] if not df.empty else 0
    if change_index_sz == 0:
        df = ts.get_realtime_quotes('sh')
        curr_price = float(df.price[0])
        pre_close = float(df.pre_close[0])
        change_index_sz = round((curr_price - pre_close)/pre_close * 100, 2)

    for x in lines:
        data = x.split(' ')
        code = data[0]
        if not filter(code):
            continue
        name = data[3]
        # 需要取得date-2,date-1,date,date+1的数据
        # 1,2,3分别对应3yang标的day1,2,3的数据
        cursor.execute(
            'select * from stock_' + code + ' where date=%s or date =%s or date =%s order by date asc' % (
            str_yestoday, str_yestoday_yestoday, str_yestoday_yestoday_yestoday))
        value = cursor.fetchall()

        change1 = float(value[0][6])
        change2 = float(value[1][6])
        change3 = float(value[2][6])
        volume1 = float(value[0][5])
        volume2 = float(value[1][5])
        volume3 = float(value[2][5])
        high1 = float(value[0][3])
        high2 = float(value[1][3])
        high3 = float(value[2][3])
        low1 = float(value[0][4])
        low2 = float(value[1][4])
        low3 = float(value[2][4])
        turnover1 = float(value[0][7])
        turnover2 = float(value[1][7])
        turnover3 = float(value[2][7])

        close1 = float(value[0][2])
        close2 = float(value[1][2])
        close3 = float(value[2][2])
        close0 = close1/(1+change1*0.01)
        high_change1 = round(((high1 - close0) / close0 * 100), 2)
        high_change2 = round((high2 - close1) / close1 * 100, 2)
        high_change3 = round((high3 - close2) / close2 * 100, 2)

        low_change1 = round((low1 - close0) / close0 * 100, 2)
        low_change2 = round((low2 - close1) / close1 * 100, 2)
        low_change3 = round((low3 - close2) / close2 * 100, 2)

        df = ts.get_realtime_quotes(code)
        curr_price = float(df.price[0])
        pre_close = float(df.pre_close[0])
        curr_high = float(df.high[0])
        curr_low = float(df.low[0])
        change4 = round((curr_price - pre_close) / pre_close * 100, 2)
        if change4 > constants.change_limit_3yang1tiao_upper_bound:
            continue

        high_change4 = round((curr_high - pre_close) / pre_close * 100, 2)
        low_change4 = round((curr_low - pre_close) / pre_close * 100, 2)
        volume4 = float(df.volume[0]) / 100
        turnover4 = round(turnover3 * volume4 / volume3, 2)

        sheet.cell(curr_row, 1).value = date
        sheet.cell(curr_row, 2).value = code
        sheet.cell(curr_row, 3).value = name

        sheet.cell(curr_row, 4).value = change1
        sheet.cell(curr_row, 5).value = change2
        sheet.cell(curr_row, 6).value = change3
        sheet.cell(curr_row, 7).value = change4
        sheet.cell(curr_row, 8).value = volume1
        sheet.cell(curr_row, 9).value = volume2
        sheet.cell(curr_row, 10).value = volume3
        sheet.cell(curr_row, 11).value = volume4
        sheet.cell(curr_row, 12).value = turnover1
        sheet.cell(curr_row, 13).value = turnover2
        sheet.cell(curr_row, 14).value = turnover3
        sheet.cell(curr_row, 15).value = turnover4
        sheet.cell(curr_row, 16).value = high_change1
        sheet.cell(curr_row, 17).value = high_change2
        sheet.cell(curr_row, 18).value = high_change3
        sheet.cell(curr_row, 19).value = high_change4
        sheet.cell(curr_row, 20).value = low_change1
        sheet.cell(curr_row, 21).value = low_change2
        sheet.cell(curr_row, 22).value = low_change3
        sheet.cell(curr_row, 23).value = low_change4
        sheet.cell(curr_row, 24).value = change_index_sz
        curr_row += 1
    sheet.delete_rows(curr_row, sheet.max_row)
    f.save(filename)

def predict_3yang1tiao():
    predict_3yang1tiao_of_sheet(constants.ml_sheet_name_predict, constants.ml_model_file_name_3yang1tiao)
    predict_3yang1tiao_of_sheet(constants.ml_sheet_name_predict_hushen, constants.ml_model_file_name_3yang1tiao_hushen)

def predict_3yang1tiao_of_sheet(sheetname, modelname):
    filename = constants.ml_data_dir + modelname
    fmodel = open(filename, 'r')
    lines = fmodel.readlines()
    fmodel.close()
    model_dict = {}
    for x in lines:
        data = x.split(' ')
        var_name = data[0]
        var_cof = data[1]
        model_dict[var_name] = float(var_cof)
    print('%s' % model_dict)

    filename = constants.ml_data_dir + constants.ml_excel_name_3yang1tiao
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]

    var_value_dict = {}
    list_result = []
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row, 2).value
        name = sheet.cell(row, 3).value

        for i in range(1, 22):
            key = 'x%d' %i
            var_value_dict[key] = sheet.cell(row, 3+i).value

        y = 0
        for key in var_value_dict:
            y += var_value_dict[key] * model_dict[key]
        y = round(y, 2)
        # print('%s %s 预测下一交易日最大涨幅为:%s%%' % (code, name, y))
        list_result.append([code, name, var_value_dict['x4'], y])
        sheet.cell(row, 25).value = y
    f.save(filename)

    list_result.sort(key=take_third, reverse=True)
    # print("%s" %list_result)
    print('\n\n')
    today = sheet.cell(2, 1).value
    filename = constants.ml_report_dir + today + '_' + sheetname + '_' + constants.ml_predict_report_filename_3yang1tiao
    fp = open(filename, 'w')
    for x in list_result:
        str = '%s %s 今日涨幅：%s%% 预测明日涨幅：%s%%' % (x[0], x[1], x[2], x[3])
        print(str)
        str+='\n'
        fp.write(str)
    fp.close()

def prepare_data():
    prepare_data_with_filter(constants.stock_filter_chuangyeban, constants.ml_sheet_name_predict)
    prepare_data_with_filter(constants.stock_filter_hushen, constants.ml_sheet_name_predict_hushen)

def prepare_data_with_filter(filter, sheetname):
    find_stock.valid_stock_2to3()

    date = time.strftime('%Y-%m-%d')
    filename = constants.report_dir + date + constants.filename_2to3
    # 遍历所有2进3标的，将所需的数据写入到excel表predict页中
    fp = open(filename, 'r')
    lines = fp.readlines()
    fp.close()

    filename = constants.ml_data_dir + constants.ml_excel_name
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]
    curr_row = 2

    now = datetime.date(*map(int, date.split('-')))
    str_yestoday, yestoday = find_stock.get_pre_trade_day(now)
    str_theday_before_yestoday, theday_before_yestoday = find_stock.get_pre_trade_day(yestoday)
    conn = mysql.connector.connect(user=constants.mysql_user, password=constants.mysql_password,
                                   database=constants.mysql_database_name)
    cursor = conn.cursor()

    df = ts.get_hist_data('sh', start=date, end=date)
    change_index_sz = df.p_change[0] if not df.empty else 0
    if change_index_sz == 0:
        df = ts.get_realtime_quotes('sh')
        curr_price = float(df.price[0])
        pre_close = float(df.pre_close[0])
        change_index_sz = round((curr_price - pre_close) / pre_close * 100, 2)

    for x in lines:
        data = x.split(' ')
        code = data[0]
        if not filter(code):
            continue
        name = data[3]
        # 需要取得date-2,date-1,date,date+1的数据
        # 1,2,3分别对应3yang标的day1,2,3的数据
        cursor.execute(
            'select * from stock_' + code + ' where date=%s or date =%s order by date asc' % (str_yestoday, str_theday_before_yestoday))  # 当天
        value = cursor.fetchall()

        change1 = float(value[0][6])
        change2 = float(value[1][6])
        volume1 = float(value[0][5])
        volume2 = float(value[1][5])
        turnover1 = float(value[0][7])
        turnover2 = float(value[1][7])
        high1 = float(value[0][3])
        high2 = float(value[1][3])
        low1 = float(value[0][4])
        low2 = float(value[1][4])
        close1 = float(value[0][2])
        close2 = float(value[1][2])
        close0 = close1 / (1 + change1 * 0.01)
        high_change1 = round(((high1 - close0) / close0 * 100), 2)
        high_change2 = round((high2 - close1) / close1 * 100, 2)
        low_change1 = round((low1 - close0) / close0 * 100, 2)
        low_change2 = round((low2 - close1) / close1 * 100, 2)

        df = ts.get_realtime_quotes(code)
        curr_price = float(df.price[0])
        pre_close = float(df.pre_close[0])
        curr_high = float(df.high[0])
        curr_low = float(df.low[0])

        change3 = round((curr_price - pre_close)/pre_close * 100, 2)
        volume3 = float(df.volume[0])/100
        turnover3 = round(turnover2 * volume3/volume2, 2)
        high_change3 = round((curr_high - pre_close)/pre_close * 100, 2)
        low_change3 = round((curr_low - pre_close)/pre_close * 100, 2)

        sheet.cell(curr_row, 1).value = date
        sheet.cell(curr_row, 2).value = code
        sheet.cell(curr_row, 3).value = name

        sheet.cell(curr_row, 4).value = change1
        sheet.cell(curr_row, 5).value = change2
        sheet.cell(curr_row, 6).value = change3
        sheet.cell(curr_row, 7).value = volume1
        sheet.cell(curr_row, 8).value = volume2
        sheet.cell(curr_row, 9).value = volume3
        sheet.cell(curr_row, 10).value = turnover1
        sheet.cell(curr_row, 11).value = turnover2
        sheet.cell(curr_row, 12).value = turnover3
        sheet.cell(curr_row, 13).value = high_change1
        sheet.cell(curr_row, 14).value = high_change2
        sheet.cell(curr_row, 15).value = high_change3
        sheet.cell(curr_row, 16).value = low_change1
        sheet.cell(curr_row, 17).value = low_change2
        sheet.cell(curr_row, 18).value = low_change3
        sheet.cell(curr_row, 19).value = change_index_sz
        curr_row += 1
    sheet.delete_rows(curr_row, sheet.max_row)
    f.save(filename)

def predict():
    predict_of_sheet(constants.ml_sheet_name_predict, constants.ml_model_file_name)
    predict_of_sheet(constants.ml_sheet_name_predict_hushen, constants.ml_model_file_name_hushen)

# 从模型中取出x1~9的系数，计算y
def predict_of_sheet(sheetname, modelname):
    filename = constants.ml_data_dir + modelname
    fmodel = open(filename, 'r')
    lines = fmodel.readlines()
    fmodel.close()
    model_dict = {}
    for x in lines:
        data = x.split(' ')
        var_name = data[0]
        var_cof = data[1]
        model_dict[var_name] = float(var_cof)
    print('%s' %model_dict)

    filename = constants.ml_data_dir + constants.ml_excel_name
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]

    var_value_dict = {}
    list_result = []
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row, 2).value
        name = sheet.cell(row, 3).value

        for i in range(1, 17):
            key = 'x%d' %i
            var_value_dict[key] = sheet.cell(row, 3+i).value

        y = 0
        for key in var_value_dict:
            y += var_value_dict[key] * model_dict[key]
        y = round(y, 2)
        # print('%s %s 预测下一交易日最大涨幅为:%s%%' %(code, name, y))
        list_result.append([code, name, var_value_dict['x3'], y])
        sheet.cell(row, 20).value = y
    f.save(filename)

    list_result.sort(key=take_third, reverse=True)
    # print("%s" %list_result)
    print('\n\n')
    today = sheet.cell(2, 1).value
    if not today:
        today = time.strftime('%Y-%m-%d')
    filename = constants.ml_report_dir + today + '_' + sheetname + '_' + constants.ml_predict_report_filename_3yang
    fp = open(filename, 'w')
    for x in list_result:
        str = '%s %s 今日涨幅：%s%% 预测明日涨幅：%s%%' % (x[0], x[1], x[2], x[3])
        print(str)
        str+='\n'
        fp.write(str)
    fp.close()

prepare_data()
predict()
print('-----------------------------------\n\n')
prepare_data_3yang1tiao()
predict_3yang1tiao()