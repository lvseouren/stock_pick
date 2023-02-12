# 准备当天的实时数据
import openpyxl

import constants


def prepare_data():
    return

# 从模型中取出x1~9的系数，计算y
def predict():
    filename = constants.ml_data_dir + constants.ml_model_file_name
    fmodel = open(filename, 'r')
    lines = fmodel.readlines()
    fmodel.close()
    model_dict = {}
    for x in lines:
        data = x.split(' ')
        var_name = data[0]
        var_cof = data[1]
        model_dict[var_name] = float(var_cof)
    print('%s' %model_dict)

    filename = constants.ml_data_dir + constants.ml_excel_name
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[constants.ml_sheet_name_predict]

    colume_start = 2
    var_value_dict = {}
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row, 2).value
        name = sheet.cell(row, 3).value
        var_value_dict['x1'] = sheet.cell(row, 4).value
        var_value_dict['x2'] = sheet.cell(row, 5).value
        var_value_dict['x3'] = sheet.cell(row, 6).value
        var_value_dict['x4'] = sheet.cell(row, 7).value
        var_value_dict['x5'] = sheet.cell(row, 8).value
        var_value_dict['x6'] = sheet.cell(row, 9).value
        var_value_dict['x7'] = sheet.cell(row, 10).value
        var_value_dict['x8'] = sheet.cell(row, 11).value
        var_value_dict['x9'] = sheet.cell(row, 12).value

        y = 0
        for key in var_value_dict:
            y += var_value_dict[key] * model_dict[key]
        y = round(y, 2)
        print('%s %s 预测下一交易日最大涨幅为:%s', code, name, y)
        sheet.cell(row, 13).value = y
    f.save(filename)
predict()