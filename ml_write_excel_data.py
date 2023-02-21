# -*- coding:utf-8 -*-
import datetime
import time

import mysql
import openpyxl
import constants
import find_stock
import tushare as ts
from MachineLearning import linear_regress

def write_to_excel(sheet, date, filter, change_index_sz):
    last_date = sheet.cell(sheet.max_row, 1).value
    if last_date == date:
        return False

    filename = constants.report_dir + date + constants.filename_3yang_list
    fp = open(filename, "r")
    lines = fp.readlines()
    fp.close()
    # 第1、2行为表头
    # 第3行开始写入数据
    # 遍历3yang标的，取得需要的数据，写入对应的xi，取得其次日的最高涨幅写入y

    now = datetime.date(*map(int, date.split('-')))
    today = constants.get_date_str_for_datebase(now)
    str_yestoday, yestoday = find_stock.get_pre_trade_day(now)
    str_theday_before_yestoday, theday_before_yestoday = find_stock.get_pre_trade_day(yestoday)
    str_next_day, next_day = find_stock.get_next_trade_day(now)
    conn = mysql.connector.connect(user=constants.mysql_user, password=constants.mysql_password,
                                   database=constants.mysql_database_name)
    cursor = conn.cursor()

    # df = ts.get_hist_data('sh', start=date, end=date)
    # change_index_sz = df.p_change[0]

    curr_row = sheet.max_row+1
    for x in lines:
        data = x.split(' ')
        code = data[0]
        if not filter(code):
            continue

        name = data[3]

        # 需要取得date-2,date-1,date,date+1的数据
        # 1,2,3分别对应3yang标的day1,2,3的数据
        cursor.execute(
            'select * from stock_' + code + ' where date=%s or date =%s or date =%s or date =%s order by date asc' % (today,
                str_yestoday, str_theday_before_yestoday, str_next_day))  # 当天
        value = cursor.fetchall()
        try:
            high1 = float(value[0][3])
            high2 = float(value[1][3])
            high3 = float(value[2][3])
            low1 = float(value[0][4])
            low2 = float(value[1][4])
            low3 = float(value[2][4])
            change1 = float(value[0][6])
            change2 = float(value[1][6])
            change3 = float(value[2][6])
            volume1 = float(value[0][5])
            volume2 = float(value[1][5])
            volume3 = float(value[2][5])
            turnover1 = float(value[0][7])
            turnover2 = float(value[1][7])
            turnover3 = float(value[2][7])
            close3 = float(value[2][2])

            close1 = float(value[0][2])
            close2 = float(value[1][2])
            close0 = close1 / (1 + change1 * 0.01)
            high_change1 = round(((high1 - close0) / close0 * 100), 2)
            high_change2 = round((high2 - close1) / close1 * 100, 2)
            high_change3 = round((high3 - close2) / close2 * 100, 2)
            low_change1 = round((low1 - close0) / close0 * 100, 2)
            low_change2 = round((low2 - close1) / close1 * 100, 2)
            low_change3 = round((low3 - close2) / close2 * 100, 2)

            high = 0
            if len(value) > 3:
                high = float(value[3][3])
            high_change = round((high - close3)/close3 * 100, 2)
            print('%s %s %s最高涨幅为:%s' %(code, name, str_next_day, high_change))

            sheet.cell(curr_row, 1).value = date
            sheet.cell(curr_row, 2).value = code
            sheet.cell(curr_row, 3).value = name

            sheet.cell(curr_row, 4).value = change1
            sheet.cell(curr_row, 5).value = change2
            sheet.cell(curr_row, 6).value = change3
            sheet.cell(curr_row, 7).value = volume1
            sheet.cell(curr_row, 8).value = volume2
            sheet.cell(curr_row, 9).value = volume3
            sheet.cell(curr_row, 10).value = turnover1
            sheet.cell(curr_row, 11).value = turnover2
            sheet.cell(curr_row, 12).value = turnover3
            sheet.cell(curr_row, 13).value = high_change1
            sheet.cell(curr_row, 14).value = high_change2
            sheet.cell(curr_row, 15).value = high_change3
            sheet.cell(curr_row, 16).value = low_change1
            sheet.cell(curr_row, 17).value = low_change2
            sheet.cell(curr_row, 18).value = low_change3
            sheet.cell(curr_row, 19).value = change_index_sz
            sheet.cell(curr_row, 20).value = high_change

            curr_row += 1
        except:
            print("%s %s %s数据异常" %(code, name, date))

    return True

def prepare_data_with_filter(starttime, endtime, sheetname, filter):
    is_dirty = False
    filename = constants.ml_data_dir + constants.ml_excel_name
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]
    df = ts.get_hist_data('sh', '2022-01-01', '2023-01-30')
    constants.cache_df = df
    constants.has_cache = True
    df = ts.get_hist_data('sh', starttime, endtime)

    try:
        for i in reversed(range(0, len(df))):
            # 获取股票日期，并转格式（这里为什么要转格式，是因为之前我2018-03-15这样的格式写入数据库的时候，通过通配符%之后他居然给我把-符号当做减号给算出来了查看数据库日期就是2000百思不得其解想了很久最后决定转换格式）
            date = df.index[i]
            # date = constants.change_date_str_format(date, '%Y%m%d', '%Y-%m-%d')
            changes = df.p_change
            change = changes[date]
            print("写入%s的数据" %date)
            if date == '2022-01-21':
                print('wtf')
            is_dirty = True if write_to_excel(sheet, date, filter, change) or is_dirty else False
    except:
        print('wtf prepare_data')
    f.save(filename)
    return is_dirty

def prepare_data(starttime, endtime):
    isDirty = False
    isDirty = prepare_data_with_filter(starttime, endtime, constants.ml_sheetname_data,
                                                  constants.stock_filter_chuangyeban)
    time.sleep(3)
    isDirty2 = prepare_data_with_filter(starttime, endtime, constants.ml_sheetname_data_hushen,
                                                   constants.stock_filter_hushen)
    return isDirty or isDirty2

def write_to_excel_3yang1tiao(sheet, date, filter, change_index_sz):
    last_date = sheet.cell(sheet.max_row, 1).value
    if last_date == date:
        return False

    filename = constants.report_dir + date + constants.filename_3yang_list
    fp = open(filename, "r")
    lines = fp.readlines()
    fp.close()
    # 第1、2行为表头
    # 第3行开始写入数据
    # 遍历3yang标的，取得需要的数据，写入对应的xi，取得其次日的最高涨幅写入y

    now = datetime.date(*map(int, date.split('-')))
    today = constants.get_date_str_for_datebase(now)
    str_yestoday, yestoday = find_stock.get_pre_trade_day(now)
    str_theday_before_yestoday, theday_before_yestoday = find_stock.get_pre_trade_day(yestoday)
    str_next_day, next_day = find_stock.get_next_trade_day(now)
    str_next_next_day, next_next_day = find_stock.get_next_trade_day(next_day)

    conn = mysql.connector.connect(user=constants.mysql_user, password=constants.mysql_password,
                                   database=constants.mysql_database_name)
    cursor = conn.cursor()
    curr_row = sheet.max_row + 1

    # df = ts.get_hist_data('sh', start=date, end=date)
    # change_index_sz = df.p_change[0]

    for x in lines:
        data = x.split(' ')
        code = data[0]
        if not filter(code):
            continue
        name = data[3]

        # 需要取得date-2,date-1,date,date+1的数据
        # 1,2,3分别对应3yang标的day1,2,3的数据
        cursor.execute(
            'select * from stock_' + code + ' where date=%s or date =%s or date =%s or date =%s or date =%s order by date asc' % (today,
            str_yestoday, str_theday_before_yestoday, str_next_day, str_next_next_day))  # 当天
        value = cursor.fetchall()

        high1 = float(value[0][3])
        high2 = float(value[1][3])
        high3 = float(value[2][3])
        high4 = float(value[3][3])
        low1 = float(value[0][4])
        low2 = float(value[1][4])
        low3 = float(value[2][4])
        low4 = float(value[3][4])
        change1 = float(value[0][6])
        change2 = float(value[1][6])
        change3 = float(value[2][6])
        change4 = float(value[3][6])
        if change4 > constants.change_limit_3yang1tiao_upper_bound:
            continue
        volume1 = float(value[0][5])
        volume2 = float(value[1][5])
        volume3 = float(value[2][5])
        volume4 = float(value[3][5])
        turnover1 = float(value[0][7])
        turnover2 = float(value[1][7])
        turnover3 = float(value[2][7])
        turnover4 = float(value[3][7])
        close4 = float(value[3][2])

        close1 = float(value[0][2])
        close2 = float(value[1][2])
        close3 = float(value[2][2])
        close0 = close1/(1+change1*0.01)
        high_change1 = round(((high1 - close0) / close0 * 100), 2)
        high_change2 = round((high2 - close1) / close1 * 100, 2)
        high_change3 = round((high3 - close2) / close2 * 100, 2)
        high_change4 = round((high4 - close3) / close3 * 100, 2)
        low_change1 = round((low1 - close0) / close0 * 100, 2)
        low_change2 = round((low2 - close1) / close1 * 100, 2)
        low_change3 = round((low3 - close2) / close2 * 100, 2)
        low_change4 = round((low4 - close3) / close3 * 100, 2)

        high = 0
        change5 = 0
        if len(value) > 4:
            high = float(value[4][3])
            change5 = float(value[4][6])

        high_change = round((high - close4) / close4 * 100, 2)
        print('%s %s %s最高涨幅为:%s' % (code, name, str_next_day, high_change))
        if code == '600970':
            print('gotcha')
        sheet.cell(curr_row, 1).value = date
        sheet.cell(curr_row, 2).value = code
        sheet.cell(curr_row, 3).value = name

        sheet.cell(curr_row, 4).value = change1
        sheet.cell(curr_row, 5).value = change2
        sheet.cell(curr_row, 6).value = change3
        sheet.cell(curr_row, 7).value = change4
        sheet.cell(curr_row, 8).value = volume1
        sheet.cell(curr_row, 9).value = volume2
        sheet.cell(curr_row, 10).value = volume3
        sheet.cell(curr_row, 11).value = volume4
        sheet.cell(curr_row, 12).value = turnover1
        sheet.cell(curr_row, 13).value = turnover2
        sheet.cell(curr_row, 14).value = turnover3
        sheet.cell(curr_row, 15).value = turnover4
        sheet.cell(curr_row, 16).value = high_change1
        sheet.cell(curr_row, 17).value = high_change2
        sheet.cell(curr_row, 18).value = high_change3
        sheet.cell(curr_row, 19).value = high_change4
        sheet.cell(curr_row, 20).value = low_change1
        sheet.cell(curr_row, 21).value = low_change2
        sheet.cell(curr_row, 22).value = low_change3
        sheet.cell(curr_row, 23).value = low_change4
        sheet.cell(curr_row, 24).value = change_index_sz
        sheet.cell(curr_row, 25).value = change5
        curr_row += 1

    return True

def prepare_data_3yang1tiao(starttime, endtime):
    isDirty = False
    isDirty = prepare_data_3yang1tiao_with_filter(starttime, endtime, constants.ml_sheetname_data,
                                        constants.stock_filter_chuangyeban)
    time.sleep(3)
    isDirty2 = prepare_data_3yang1tiao_with_filter(starttime, endtime, constants.ml_sheetname_data_hushen,
                                        constants.stock_filter_hushen)
    return isDirty or isDirty2

def prepare_data_3yang1tiao_with_filter(starttime, endtime, sheetname, filter):
    is_dirty = False
    filename = constants.ml_data_dir + constants.ml_excel_name_3yang1tiao
    print(filename)
    f = openpyxl.open(filename)
    sheet = f[sheetname]
    df = ts.get_hist_data('sh', starttime, endtime)
    try:
        for i in reversed(range(0, len(df))):
            date = df.index[i]
            # date = constants.change_date_str_format(date, '%Y%m%d', '%Y-%m-%d')
            changes = df.p_change
            change = changes[date]
            print("写入%s的数据" % date)
            
            temp_dirty = write_to_excel_3yang1tiao(sheet, date, filter, change)
            is_dirty = is_dirty or temp_dirty
            time.sleep(1)
    except:
        print('wtf prepare_data_3yang1tiao_with_filter')
    f.save(filename)
    return is_dirty

prepare_data('2022-01-10','2023-01-13')
# linear_regress.mul_lr_3yang()
prepare_data_3yang1tiao('2022-01-10','2023-01-13')
# linear_regress.mul_lr_3yang1tiao()



